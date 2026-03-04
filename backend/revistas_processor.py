import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def split_pipe(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.lower() in ('nan', 'none'):
        return []
    return [v.strip() for v in s.split('|') if v.strip()]


def process_revistas_excel_to_json(excel_path: str, output_path: str):
    print(f"Reading Excel file: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    def get_named(name, row):
        idx = col_idx.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    revistas = []
    publishers = set()
    enfoques = set()
    disciplines = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        journal = get_named('Journal', row)
        if not journal or str(journal).strip() == '':
            continue

        publisher = str(get_named('Publisher', row) or '').strip()
        enfoque = str(get_named('Enfoque principal (datos/base de datos)', row) or '').strip()
        disciplinas = split_pipe(get_named('Disciplinas principales', row))

        if publisher:
            publishers.add(publisher)
        if enfoque:
            enfoques.add(enfoque)
        for d in disciplinas:
            disciplines.add(d)

        revista = {
            'id': idx,
            'journal': str(journal).strip(),
            'issn': str(get_named('ISSN (impreso)', row) or '').strip() or None,
            'eissn': str(get_named('eISSN', row) or '').strip() or None,
            'publisher': publisher or None,
            'enfoque': enfoque or None,
            'disciplinas': disciplinas,
            'sitioWeb': str(get_named('Sitio web de la revista', row) or '').strip() or None,
        }

        revistas.append(revista)

    output_data = {
        'metadata': {
            'totalRevistas': len(revistas),
            'publishers': len(publishers),
            'disciplinas': len(disciplines),
            'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'version': '1.0'
        },
        'revistas': revistas,
        'facets': {
            'publishers': sorted([p for p in publishers if p]),
            'enfoques': sorted([e for e in enfoques if e]),
            'disciplinas': sorted([d for d in disciplines if d]),
        }
    }

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"Successfully created: {output_path}")
    return output_data


if __name__ == '__main__':
    excel_file = 'List_revista.xlsx'
    json_file = 'public/revistas.json'

    process_revistas_excel_to_json(excel_file, json_file)
