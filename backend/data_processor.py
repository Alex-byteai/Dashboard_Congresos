import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd # Use pandas for easier merging/lookup if available, or stick to openpyxl if dep is issue. 
# actually, existing code uses openpyxl. sticking to it for consistency unless pandas is better for this join.
# The user env has pandas installed (implied by previous context). I'll use pandas for the join file as it's cleaner.



def parse_date(date_val):
    """Parse various date formats to ISO format.
    
    IMPORTANT: When openpyxl returns a datetime object, it means Excel stored
    the cell as a date. However, dates entered as DD/MM/YYYY text get interpreted
    by Excel (US locale) as MM/DD/YYYY, so day and month are SWAPPED.
    Text strings with '/' are parsed correctly as DD/MM/YYYY.
    """
    if date_val is None or date_val == '':
        return None

    # If openpyxl returned a datetime object: Excel swapped day/month on entry.
    # Swap them back to get the correct date.
    if isinstance(date_val, datetime):
        # Swap: stored month is actually the day, stored day is actually the month
        corrected = date_val.replace(day=date_val.month, month=date_val.day)
        return corrected.strftime('%Y-%m-%d')

    date_str = str(date_val).strip()

    # Text in DD/MM/YYYY or D/M/YYYY format (correct as-is)
    if '/' in date_str:
        try:
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        except:
            pass

    # Already in YYYY-MM-DD format
    if '-' in date_str and len(date_str) >= 10:
        return date_str[:10]

    return date_str

def get_deadline_status(deadline_str):
    """Calculate deadline urgency status"""
    if not deadline_str:
        return 'unknown'
    
    try:
        # Parse the date
        if isinstance(deadline_str, datetime):
            deadline_date = deadline_str
        else:
            # Try to parse ISO format
            deadline_date = datetime.fromisoformat(deadline_str)
        
        today = datetime.now()
        days_until = (deadline_date - today).days
        
        if days_until < 0:
            return 'passed'
        elif days_until <= 30:
            return 'urgent'
        elif days_until <= 90:
            return 'upcoming'
        else:
            return 'future'
    except:
        return 'unknown'



# Full taxonomy: Categoría → Línea → [Sublíneas]
# This is the authoritative hierarchy defined by ULIMA.
FULL_TAXONOMY = {
    'INNOVACIÓN Y TECNOLOGÍA DIGITAL': {
        'Inteligencia artificial y computación avanzada': [
            'Machine learning y deep learning',
            'Procesamiento de lenguaje natural',
            'Visión computacional',
            'Sistemas autónomos y robótica',
        ],
        'Transformación digital': [
            'Tecnologías emergentes',
            'Ciberseguridad y privacidad',
            'Internet de las cosas (IoT)',
            'Computación cuántica',
            'Diseño y construcción virtual',
        ],
        'Experiencia digital humana': [
            'Interacción humano-computadora',
            'Realidad virtual y aumentada',
            'Diseño de interfaces adaptativas',
        ],
    },
    'DESARROLLO SOSTENIBLE Y MEDIOAMBIENTE': {
        'Sostenibilidad y cambio climático': [
            'Energías renovables',
            'Economía circular',
            'Gestión sostenible de recursos',
            'Adaptación al cambio climático',
        ],
        'Ciudades inteligentes y sostenibles': [
            'Urbanismo sostenible',
            'Movilidad urbana',
            'Infraestructura sostenible',
            'Gestión inteligente de recursos',
        ],
        'Tecnología y ecosistemas': [
            'Tecnologías limpias',
            'Biodiversidad y conservación',
            'Gestión de residuos',
            'Materiales avanzados',
        ],
    },
    'SOCIEDAD Y COMPORTAMIENTO HUMANO': {
        'Bienestar y desarrollo humano': [
            'Salud mental y bienestar',
            'Educación, desarrollo cognitivo y socioafectivo',
            'Comportamiento social',
            'Mujer, cultura y sociedad',
            'Pobreza e informalidad',
        ],
        'Comunicación y cultura digital': [
            'Medios digitales y sociedad',
            'Comunicación intercultural',
            'Narrativas transmedia',
            'Comportamiento digital',
        ],
        'Ética, gobernanza y responsabilidad social': [
            'Ética y gobernanza',
            'Responsabilidad social',
            'Derechos humanos y tecnología',
        ],
    },
    'GESTIÓN Y ECONOMÍA DEL CONOCIMIENTO': {
        'Innovación empresarial': [
            'Modelos de negocio digitales',
            'Emprendimiento tecnológico',
            'Gestión de la innovación',
            'Transformación organizacional',
        ],
        'Economía digital': [
            'Fintech y servicios financieros',
            'Mercados globales',
            'Análisis de datos económicos',
            'Economía de plataformas',
        ],
        'Gestión del conocimiento': [
            'Gestión del capital intelectual',
            'Aprendizaje organizacional',
            'Transferencia de conocimiento',
            'Inteligencia de negocios',
        ],
    },
}

def split_semicolon(value):
    """Split a semicolon-separated string into a sorted list of stripped values"""
    if not value or str(value).strip() in ('', 'nan', 'None'):
        return []
    return sorted([v.strip() for v in str(value).split(';') if v.strip()])

def process_excel_to_json(excel_path, output_path):
    """Main processing function"""
    print(f"Reading Excel file: {excel_path}")
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"Found {len(headers)} columns")
    
    congresses = []
    countries = set()
    modalities = {}

    # Build a column-name → index map from the header row (robust to column changes)
    col_idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}
    print(f"Column map: {col_idx}")

    def get_named(name, row):
        """Get a cell value by column name, returns None if column not found."""
        idx = col_idx.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    # Process each row (skip header)
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):

        evento  = get_named('Evento', row)
        nombre  = get_named('Nombre Completo', row)

        # Skip empty rows
        if not evento and not nombre:
            continue

        # Parse dates
        fecha_inicio = parse_date(get_named('Fecha inicio', row))
        fecha_fin    = parse_date(get_named('Fecha fin', row))
        deadline     = parse_date(get_named('Deadline', row))

        ciudad   = str(get_named('Ciudad', row) or '')
        pais     = str(get_named('Pais', row) or '')
        modalidad = str(get_named('Modalidad', row) or '')

        # Track statistics
        if pais: countries.add(pais)
        if modalidad: modalities[modalidad] = modalities.get(modalidad, 0) + 1

        # Parse 3-level hierarchy (semicolon-separated values) by column name
        categorias = split_semicolon(get_named('categoria_ulima', row))
        lineas     = split_semicolon(get_named('linea_ulima', row))
        sublineas  = split_semicolon(get_named('sublinea_ulima', row))

        congress = {
            'id': idx,
            'evento': str(evento or ''),
            'nombreCompleto': str(nombre or ''),
            'disciplina': str(get_named('Disciplina', row) or ''),
            'categoria': categorias,
            'linea': lineas,
            'sublinea': sublineas,
            'fechaInicio': fecha_inicio,
            'fechaFin': fecha_fin,
            'lugar': str(get_named('Lugar', row) or ''),
            'ciudad': ciudad,
            'pais': pais,
            'modalidad': modalidad,
            'deadline': deadline,
            'deadlineStatus': get_deadline_status(deadline),
            'publicacion': str(get_named('Publicación', row) or ''),
            'enlace': str(get_named('Enlace', row) or '')
        }

        congresses.append(congress)

    
    print(f"Processed {len(congresses)} congresses")
    print(f"Using full taxonomy: {len(FULL_TAXONOMY)} categorias")
    
    # Create output structure
    output_data = {
        'metadata': {
            'totalCongresses': len(congresses),
            'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'version': '2.1'
        },
        'taxonomy': FULL_TAXONOMY,
        'congresses': congresses
    }
    
    # Write JSON
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"Successfully created: {output_path}")
    print(f"Statistics:")
    print(f"   - Total congresses: {len(congresses)}")
    print(f"   - Countries: {len(countries)}")
    print(f"   - Modalities: {modalities}")
    
    return output_data

if __name__ == '__main__':
    excel_file = 'List_congreso.xlsx'
    json_file = 'public/congresses.json'
    
    try:
        process_excel_to_json(excel_file, json_file)
        print("\nData processing complete!")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
